from utils import checkNumber
from openpyxl import load_workbook


def get():
    doc = load_workbook(filename='index.xlsx')

    ws = doc.active

    a = []  # supplier
    b = []  # wholesaler
    e = []  # retailer
    f = []
    g = []
    i = []
    ix = []
    j = []  # collection-center
    hj = []
    m = []  # material
    p = []  # product
    t = []
    v = []

    for _row in range(2, ws.max_row+1):
        _indxname = ws.cell(_row, 1).value
        _value = ws.cell(_row, 2).value
        if checkNumber(_value):

            if (str(_indxname) == 'a'):
                for _ in range(_value):
                    a.append(_+1)

            elif (str(_indxname) == 'b'):
                for _ in range(_value):
                    b.append(_+1)

            elif (str(_indxname) == 'e'):
                for _ in range(_value):
                    e.append(_+1)

            elif (str(_indxname) == 'f'):
                for _ in range(_value):
                    f.append(_+1)

            elif (str(_indxname) == 'g'):
                for _ in range(_value):
                    g.append(_+1)

            elif (str(_indxname) == 'i'):
                for _ in range(_value):
                    i.append(_+1)

            elif (str(_indxname) == 'ix'):
                for _ in range(_value):
                    ix.append(_+1)

            elif (str(_indxname) == 'j'):
                for _ in range(_value):
                    j.append(_+1)

            elif (str(_indxname) == 'hj'):
                for _ in range(_value):
                    hj.append(_+1)

            elif (str(_indxname) == 'm'):
                for _ in range(_value):
                    m.append(_+1)

            elif (str(_indxname) == 'p'):
                for _ in range(_value):
                    p.append(_+1)

            elif (str(_indxname) == 't'):
                for _ in range(_value):
                    t.append(_+1)

            elif (str(_indxname) == 'v'):
                for _ in range(_value):
                    v.append(_+1)

    return a, b, e, f, g, i, ix, j, hj, m, p, t, v
import utils
from openpyxl import load_workbook


def get():
    doc = load_workbook(filename='normal-table.xlsx')

    # 3 level nesting
    sn = doc.sheetnames
    # data = []
    T = []
    Tx = []
    Txx = []
    ct = []
    cs = []
    cf = []
    cpxx = []
    alpha = []
    I = []
    Ix = []
    ch = []
    Ts = []
    Tsx = []
    ck = []
    fx = []
    czx = []
    Sp = []
    chx = []
    Spx = []
    cab = []
    BR = []
    c = []
    cx = []
    cxx = []
    vol = []
    volx = []
    volxx = []

    for sheet in sn:
        ws = doc[str(sheet)]
        # print(sheet)

        # init array
        _data = [[0 for _ in range(2, ws.max_column+1)]
                 for _ in range(2, ws.max_row+1)]

        # assign values
        for i in range(2, ws.max_row+1):
            for j in range(2, ws.max_column+1):
                _data[i-2][j-2] = ws.cell(i, j).value
        # data.append(_data)

        fixd_d = utils.convert_to_tuple(_data)
        if (str(sheet) == 'T'):
            T = fixd_d
        elif (str(sheet) == 'Tx'):
            Tx = fixd_d
        elif (str(sheet) == 'Txx'):
            Txx = fixd_d
        elif (str(sheet) == 'ct'):
            ct = fixd_d
        elif (str(sheet) == 'cs'):
            cs = fixd_d
        elif (str(sheet) == 'cf'):
            cf = fixd_d
        elif (str(sheet) == 'cpxx'):
            cpxx = fixd_d
        elif (str(sheet) == 'alpha'):
            alpha = fixd_d
        elif (str(sheet) == 'I'):
            I = fixd_d
        elif (str(sheet) == 'Ix'):
            Ix = fixd_d
        elif (str(sheet) == 'ch'):
            ch = fixd_d
        elif (str(sheet) == 'Ts'):
            Ts = fixd_d
        elif (str(sheet) == 'Tsx'):
            Tsx = fixd_d
        elif (str(sheet) == 'ck'):
            ck = fixd_d
        elif (str(sheet) == 'fx'):
            fx = fixd_d
        elif (str(sheet) == 'czx'):
            czx = fixd_d
        elif (str(sheet) == 'Sp'):
            Sp = fixd_d
        elif (str(sheet) == 'chx'):
            chx = fixd_d
        elif (str(sheet) == 'Spx'):
            Spx = fixd_d
        elif (str(sheet) == 'cab'):
            cab = fixd_d
        elif (str(sheet) == 'BR'):
            BR = fixd_d
        elif (str(sheet) == 'c'):
            c = fixd_d
        elif (str(sheet) == 'cx'):
            cx = fixd_d
        elif (str(sheet) == 'cxx'):
            cxx = fixd_d
        elif (str(sheet) == 'vol'):
            vol = fixd_d
        elif (str(sheet) == 'volx'):
            volx = fixd_d
        elif (str(sheet) == 'volxx'):
            volxx = fixd_d

    # d = utils.convert_to_tuple(data)
    # print(d)
    return T, Tx, Txx, ct, cs, cf, cpxx, alpha, I, Ix, ch, Ts, Tsx, ck, fx, czx, Sp, chx, Spx, cab, BR, c, cx, cxx, vol, volx, volxx
    # T
    # Tx
    # Txx
    # ct
    # cs
    # cf
    # cpxx
    # alpha
    # I
    # Ix
    # ch
    # Ts
    # Tsx
    # ck
    # fx
    # czx
    # Sp
    # chx
    # Spx
    # cab
    # BR
    # c
    # cx
    # cxx
    # vol
    # volx
    # volxx
    # ([[1], [2]], [[3], [4]], [[6, 7], [5, 6]], [[5], [10]], [[100, 100], [100, 100]], [[2, 1], [3, 2]], [[5000], [6000]], [[2, 1], [3, 2]], [[5], [3]], [[6], [5]], [[20, 22], [21, 23]], [[3, 2], [2, 3]], [[3], [4]], [[10, 20], [11, 22]], [[60], [60]], [[7, 8], [6, 7]], [[100], [90], [120], [100]], [[5, 3], [4, 2]], [[80], [90], [80], [100]], [[120], [110]], [[3, 3], [2, 2]], [[15, 16], [16, 17]], [[10, 10, 10, 10], [10, 10, 10, 10]], [[2, 1], [1, 3]], [[170], [180], [160], [170]], [[75], [85]], [[40], [20]])

# print(get()) # (([1], [2]), ([3], [4]), ([6, 7], [5, 6]), ([5], [10]), ([100, 100], [100, 100]), ([2, 1], [3, 2]), ([5000], [6000]), ([2, 1], [3, 2]), ([5], [3]), ([6], [5]), ([20, 22], [21, 23]), ([3, 2], [2, 3]), ([3], [4]), ([10, 20], [11, 22]), ([60], [60]), ([7, 8], [6, 7]), ([100], [90], [120], [100]), ([5, 3], [4, 2]), ([80], [90], [80], [100]), ([120], [110]), ([3, 3], [2, 2]), ([15, 16], [16, 17]), ([10, 10, 10, 10], [10, 10, 10, 10]), ([2, 1], [1, 3]), ([170], [180], [160], [170]), ([75], [85]), ([40], [20]))
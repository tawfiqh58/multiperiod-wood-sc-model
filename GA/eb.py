from utils import convert_to_tuple
from openpyxl import load_workbook


def get():
    doc = load_workbook(filename='eb.xlsx')

    ws = doc.active
    data = []

    for i in range(2, ws.max_row+1):
        for j in range(2, ws.max_column+1):
            val = ws.cell(i, j).value

            if isinstance(val, (int, float)):
                # number; means single input
                # add as an array
                data.append([val])
            else:
                # sting input ',' saperator
                x = val.split(",")
                va = []
                for _x in range(len(x)):
                    # convert to int
                    _int = int(x[_x])
                    va.append(_int)

                data.append(va)

    d = convert_to_tuple(data)
    return d


print(get())

from openpyxl import load_workbook
fxd_cost_veh_whol_cust = load_workbook(filename='fxd_cost_veh_whol_cust.xlsx')
whol_names = fxd_cost_veh_whol_cust.sheetnames
data = []
for whol in whol_names:
    # print(whol)
    ws = fxd_cost_veh_whol_cust[str(whol)]

    # init array size
    _data = [[0 for _ in range(1, ws.max_column)]
        for _ in range(1, ws.max_row)]
    print(data)

    # assign value
    for i in range(1, ws.max_row):
        for j in range(1, ws.max_column):
            _data[i][j] = ws.cell(i+1, j+1).value
    # print(_data)
    data.append(_data)

# for whol in whol_names:
#     # print(whol)
#     ws = fxd_cost_veh_whol_cust[str(whol)]

#     # init array size
#     _data = [[0 for _ in range(1, ws.max_column+1)]
#         for _ in range(1, ws.max_row+1)]
#     # print(data)

#     # assign value
#     for i in range(0, ws.max_row):
#         for j in range(0, ws.max_column):
#             _data[i][j] = ws.cell(i+1, j+1).value
#     # print(_data)
#     data.append(_data)

# print(data)

# suppliyer
#   vehicle
#       retailer

# array length
# supplier_len = len(data)
define = []
for s in data:
    # print(s)
    for v in s:
        # print(r)
        for r in v:
            print(r)
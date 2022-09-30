import random
from openpyxl import load_workbook
fxd_cost_veh_whol_cust = load_workbook(filename='fxd_cost_veh_whol_cust.xlsx')
whol_names = fxd_cost_veh_whol_cust.sheetnames
data = []

# understanding
# for whol in whol_names:
#     # print(whol)
#     ws = fxd_cost_veh_whol_cust[str(whol)]

#     print('')
#     print(ws.max_column) # 3
#     for i in range(0, ws.max_column):
#         print(i) # 0 1 2
#     for i in range(1, ws.max_column+1):
#         # useful for cell(row=1,column=i)
#         print(i) # 1 2 3

# all rows and column
# for whol in whol_names:
#     # print(whol)
#     ws = fxd_cost_veh_whol_cust[str(whol)]

#     # init array size
#     _data = [[0 for _ in range(1, ws.max_column+1)]
#         for _ in range(1, ws.max_row+1)]
#     # print(data)

#     # assign value
#     for i in range(0, ws.max_row):
#         for j in range(0, ws.max_column):
#             _data[i][j] = ws.cell(i+1, j+1).value
#     # print(_data)
#     data.append(_data)

# exclude 1st row 1st column
for whol in whol_names:
    # print(whol)
    ws = fxd_cost_veh_whol_cust[str(whol)]

    # init array size
    _data = [[0 for _ in range(2, ws.max_column+1)]
        for _ in range(2, ws.max_row+1)]
    # print(data)

    # assign value
    for i in range(2, ws.max_row+1):
        for j in range(2, ws.max_column+1):
            _data[i-2][j-2] = ws.cell(i, j).value
    # print(_data)
    data.append(_data)

# print(data)
# print(data[2][1][2]) # get value

# index
supplier_idx = [0,1,2]
retailer_idx = [0,1]
vehicle_idx = [0,1]

a_idx = [0,1,2]
v_idx = [0,1]
t_idx = [0,1]

# index size
supplier_size = len(supplier_idx)
retailer_size = len(retailer_idx)
vehicle_size = len(vehicle_idx)
a_size = len(a_idx)
v_size = len(v_idx)
t_size = len(t_idx)

# assign random value for solution variable
w_avt = [[[0 for x in range(t_size)] for y in range(v_size)] for z in range(a_size)]
for a in range(a_size):
    for v in range(v_size):
        for t in range(t_size):
            w_avt[a][v][t] = random.randint(10,  900) # posible range of that variable 

print(w_avt)

# suppliyer
#   retailer
#       vehicle
# for s in range(supplier_size):
#     for r in range(retailer_size):
#         for v in range(vehicle_size):
#             print(data[s][r][v])
sum = 0
for a in range(a_size):
    for v in range(v_size):
        for t in range(t_size):
            
            for s in range(supplier_size):
                for r in range(retailer_size):
                    for vh in range(vehicle_size):
                        sum +=data[s][r][vh]*w_avt[a][v][t]
            
            for vh in range(vehicle_size):
                # this size is also looping each time
                print('understand it')
                # it actually will loop everytime
                # but do i need this each loop
                # or what is the use case of this loop

                # if the for loop is matched then you
                # can add those code here

            
print('sum:',sum)

# array length
# supplier_len = len(data)
# define = []
# for s in data:
#     # print(s)
#     for v in s:
#         # print(r)
#         for r in v:
#             print(r)
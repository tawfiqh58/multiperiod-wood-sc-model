from utils import convert_to_tuple
from openpyxl import load_workbook

def get():
    doc = load_workbook(filename='cd.xlsx')

    # 3 level nesting
    sn = doc.sheetnames
    data = []

    for sheet in sn:
        ws = doc[str(sheet)]

        # init array
        _data = [[0 for _ in range(2, ws.max_column+1)]
            for _ in range(2, ws.max_row+1)]

        # assign values
        for i in range(2, ws.max_row+1):
            for j in range(2, ws.max_column+1):
                _data[i-2][j-2] = ws.cell(i, j).value
        data.append(_data)
    
    d = convert_to_tuple(data)
    # print(d) # ([[1, 2], [3, 4]], [[2, 3], [4, 5]])
    # print(type(utils.convert_to_tuple(data)))
    return d
import utils
from openpyxl import load_workbook
doc = load_workbook(filename='dem.xlsx')

# 3 level nesting
sn = doc.sheetnames
data = []

for sheet in sn:
    ws = doc[str(sheet)]

    # init array
    _data = [[0 for _ in range(2, ws.max_column+1)]
        for _ in range(2, ws.max_row+1)]

    # assign values
    for i in range(2, ws.max_row+1):
        for j in range(2, ws.max_column+1):
            _data[i-2][j-2] = ws.cell(i, j).value
    data.append(_data)
d = utils.convert_to_tuple(data)
print(d) # ([[400, 400], [450, 450]], [[500, 500], [550, 550]])
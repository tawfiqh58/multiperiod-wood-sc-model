import random
from openpyxl import load_workbook
from cd import get as _cd
from chxx import get as _chxx
from chxxx import get as _chxxx
from constants import get as _constants
from cr import get as _cr
from cz import get as _cz
from dem import get as _dem
from f import get as _f
from normal_table import get as _normal_table
from P import get as _P
from utils import convert_to_tuple

# create random solution
def random_population(index):
    population=[]
    for truck in range(len(index)):
        bound=[0,10] # take bound from user
        population.append(random.randint(bound[0],  bound[1]))
    return population



def input_params():
    fxd_cost_veh_whol_cust = load_workbook(filename='fxd_cost_veh_whol_cust.xlsx')
    whol_names = fxd_cost_veh_whol_cust.sheetnames
    data = []
    # loop each sheet
    for whol in whol_names:
        # sheet 1 , 2 , 3 ..
        ws = fxd_cost_veh_whol_cust[str(whol)]

        # init array
        _data = [[0 for _ in range(2, ws.max_column+1)] for _ in range(2, ws.max_row+1)]

        # assign values
        for i in range(2, ws.max_row+1):
            for j in range(2, ws.max_column+1):
                _data[i-2][j-2] = ws.cell(i, j).value
        data.append(_data)
    
    return data

def fitness_function(x, y):
    z=0
    cd = _cd()
    chxx = _chxx()
    chxxx = _chxxx()
    cr = _cr()
    cz = _cz()
    dem = _dem()
    f = _f()
    P = _P()
    constants = _constants()
    normal_table = _normal_table()

    for index in range(len(x)):
        # have to match dimention
        z += x[index]*y[0][0][index]

    if(z==0):
        return 9999999
    return 1/z

# index_of_supplier_vehicle = [0,1]
# params = input_params()
# pop = random_population(index_of_supplier_vehicle)
# print(pop)
# print(params)
# z_value = fitness_function(pop, params)
# print(z_value)

def get_pop():
    index_of_supplier_vehicle = [0,1]
    params = input_params()
    pop = random_population(index_of_supplier_vehicle)
    # print(pop)
    # print(params)
    z_value = fitness_function(pop, params)
    print(z_value)
    return z_value, pop

def ga():
    p1 = get_pop()
    p2 = get_pop()
    
    score=0
    soln=[]
    if(p2[0]>p1[0]):
        score=p2[0]
        soln=p2[1]
    else:
        score=p1[0]
        soln=p1[1]
    
    print(score) # which has more fitness value
    print(soln)

ga()
from openpyxl import load_workbook

def get():
    doc = load_workbook(filename='constants.xlsx')
    ws = doc.active

    cinv = 0
    Cinvx = 0
    M = 0
    BigM = 0
    Hd = 0
    cp = 0
    cpx = 0

    # assign values
    for i in range(2, ws.max_row+1):

        if (str(ws.cell(i, 1).value) == 'cinv'):
            cinv = ws.cell(i, 2).value
        if (str(ws.cell(i, 1).value) == 'Cinvx'):
            Cinvx = ws.cell(i, 2).value
        if (str(ws.cell(i, 1).value) == 'M'):
            M = ws.cell(i, 2).value
        if (str(ws.cell(i, 1).value) == 'BigM'):
            BigM = ws.cell(i, 2).value
        if (str(ws.cell(i, 1).value) == 'Hd'):
            Hd = ws.cell(i, 2).value
        if (str(ws.cell(i, 1).value) == 'cp'):
            cp = ws.cell(i, 2).value
        if (str(ws.cell(i, 1).value) == 'cpx'):
            cpx = ws.cell(i, 2).value

    return cinv, Cinvx, M, BigM, Hd, cp, cpx

# print(get()) # (0, 100, 10000000, 1000000000, 12, 8000, 8000)
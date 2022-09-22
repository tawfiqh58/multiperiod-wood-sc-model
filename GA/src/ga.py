import random
from openpyxl import load_workbook

def main():
    global c1
    global A_size, Va_size, T_size

    # TODO: count length from sheet
    A_size = 2
    Va_size = 2
    T_size = 2
    
    ## c1
    #
    # 💡fixed cost can be taken statically but then how will I set array size and length
    # just depends on the index length
    #
    fxd_cost_veh_supp_fac = load_workbook(filename='fxd_cost_veh_supp_fac.xlsx')
    vsf = fxd_cost_veh_supp_fac.active
    # init array
    c1 = [[0 for _ in range(2, vsf.max_column+1)]
        for _ in range(2, vsf.max_row+1)]
    # assign value
    for i in range(2, vsf.max_row+1):
        for j in range(2, vsf.max_column+1):
            c1[i-2][j-2] = vsf.cell(i, j).value
    
    print(c1)

def fitness_test(set):
    # set size = 21
    # set will be a tuple of variables
    # variable will be the solution combination
    # decoded from bitstring

    # init pop
    # tuple (random value)
    # ...
    # tuple (cromosome as a tuple)
    # fitness test
    ## encode
    ## cross
    ## mut
    ## decode
    # tuple (cromosome as a tuple)
    # ...

    x_avt = set[11] # 12th position variable for x_avt
    # variable type
    # 1. binary [0,1]
    # 2. positive number [0,inf]

    # constraints
    panalty_cost = 0
    # TODO: add all constraints
    # if anyone of them not meet then add panalty for that soln

    # constraints #1
    # constraints #2
    # constraints #3
    # constraints #4
    # constraints #5
    # constraints #6
    # constraints #7
    # constraints #8
    # constraints #9
    # constraints #10
    # constraints #11
    # constraints #12
    # constraints #13
    # constraints #14
    # constraints #15
    # constraints #16
    # constraints #17
    # constraints #18
    # constraints #19
    # constraints #20
    # constraints #21
    # constraints #22
    # constraints #23

    # z function sections
    trans_cost = 0
    
    for a in range(len(A_size)):
        for v in range(len(Va_size)):
            for t in range(len(T_size)):
                trans_cost += c1[a][v] * x_avt[a][v][t]
    print('transportation cost: ', trans_cost)

    purchas_cost = 10
    prod_cost = 10
    maintain_cost = 10
    shortage_cost = 10
    environ_cost = 10

    # objective function
    z = trans_cost + purchas_cost + prod_cost + maintain_cost + shortage_cost + environ_cost + panalty_cost

    return z

main()
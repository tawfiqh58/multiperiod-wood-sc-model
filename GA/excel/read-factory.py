from openpyxl import load_workbook
wb = load_workbook(filename='factory.xlsx')
sheet1 = wb.active
print('rows', sheet1.max_row)  # 9
print('column', sheet1.max_column)  # 5

for i in range(1, sheet1.max_row+1):
    for j in range(1, sheet1.max_column+1):
        # Read value
        print(sheet1.cell(i, j).value)

# Add value
# sheet1.cell(row=sheet1.max_row+1, column=sheet1.max_column +
#             1, value='New value assign')

sheet1_rows = sheet1.max_row
for j in range(1, sheet1.max_column):
    # Add value
    sheet1.cell(row=sheet1_rows, column=j, value='New value: ' + str(j))

# for row in sheet1.rows:
#     for cell in row:
#         print(cell.value)

wb.save('new_saved.xlsx')

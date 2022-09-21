from email import header
from openpyxl import load_workbook
wb = load_workbook(filename='large_file.xlsx')
# print(wb.sheetnames)
ws = wb.active
print(ws)

# for row in ws.rows:
#     # print(row)
#     for cell in row:
#         # print(cell)
#         print(cell.value)

# # Close the workbook after reading
# wb.close()

# ---

# ws=wb.active

# max_row=ws.max_row
# max_column=ws.max_column

# # iterate over all cells
# for i in range(1, max_row+1):
#      for j in range(1, max_column+1):
#           # get particular cell value    
#           cell_obj=ws.cell(row=i, column=j)
#           # print cell value     
#           print(cell_obj.value, end=' | ')

# ---

# rows = ws.rows

# headers = [cell.value for cell in next(rows)]

# all_rows = []

# for row in rows:
#     data = {}
#     for title, cell in zip(headers, row):
#         data[title] = cell.value

#     all_rows.append(data)

# print(all_rows)

# ---

sheet1 = wb['Factory']

print(sheet1['A2'].value)
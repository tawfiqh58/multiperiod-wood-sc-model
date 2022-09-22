from openpyxl import load_workbook
wb = load_workbook(filename='vehicle_sending_cost.xlsx')
sheet1 = wb.active

print('rows', sheet1.max_row)  # 5
print('column', sheet1.max_column)  # 4

# init parameters
data = [[0 for _ in range(1, sheet1.max_column+1)]
        for _ in range(1, sheet1.max_row+1)]
print(data)

# assign parameter values
for i in range(0, sheet1.max_row):
    for j in range(0, sheet1.max_column):
        data[i][j] = sheet1.cell(i+1, j+1).value

# print value/work with this
print(data)

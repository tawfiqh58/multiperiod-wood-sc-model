from utils import convert_to_tuple
from openpyxl import load_workbook


def get(g):
    doc = load_workbook(filename='Gb.xlsx')

    ws = doc.active
    data = []

    for i in range(2, ws.max_row+1):
        for j in range(2, ws.max_column+1):
            val = ws.cell(i, j).value

            if isinstance(val, (int, float)):
                # number; means single input
                # add as an array
                va = []
                for _ in range(len(g)):
                    
                    if val-1 == _:
                        va.append(1)
                    else:
                        va.append(0)
                data.append(va)
            else:
                # sting input ',' saperator
                x = val.split(",")
                va = []
                for _ in range(len(g)):
                    # convert to int
                    is_matched = 0
                    for _x in range(len(x)):
                        # convert to int
                        _int = int(x[_x])

                        if _int-1 == _:
                            is_matched = 1
                    va.append(is_matched)
                data.append(va)

    d = convert_to_tuple(data)
    return d